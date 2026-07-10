"""
НейроПост v2.0 — Модуль 2: Генератор контент-плана
"""
import json
import os
import sys
from datetime import date, timedelta
from dotenv import load_dotenv
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

PASSPORT_FILE = "brand_passport.json"
OUTPUT_DIR = "output"

DAYS_RU = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

POST_TYPES_POOL = [
    "полезный пост", "история/сторителлинг", "кейс клиента",
    "анонс/оффер", "полезный пост", "вирусный/провокационный", "личное/закулисье",
]


def load_passport():
    if not os.path.exists(PASSPORT_FILE):
        print("\n❌ Паспорт Бренда не найден. Сначала запусти: python setup.py\n")
        sys.exit(1)
    with open(PASSPORT_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def generate_topics(passport, posts_per_week, weeks=2):
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        print("\n❌ GEMINI_API_KEY не найден в .env файле\n")
        sys.exit(1)

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-1.5-flash")
    total_posts = posts_per_week * weeks

    prompt = f"""Ты — контент-стратег. Создай разнообразный контент-план.
Ниша эксперта: {passport.get('who_i_am', '')}
Клиент: {passport.get('target_client', '')}
Тон: {passport.get('tone', '')}
Оффер: {passport.get('offer', '')}

Придумай {total_posts} тем для постов на 2 недели.
Чередуй типы: полезный, история, кейс, анонс, вирусный, личное.

Ответь ТОЛЬКО списком в формате (одна тема на строку):
ТИП|ТЕМА|ХЭШТЕГИ

Пример:
полезный|5 ошибок при настройке таргета|#таргет #реклама #маркетинг #ошибки #бизнес
история|Как я потеряла клиента из-за одного слова|#история #опыт #продажи #урок #личное

Дай ровно {total_posts} строк, без нумерации, без лишнего текста."""

    print(f"\n⏳ Генерирую {total_posts} тем для контент-плана...")

    try:
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        print(f"\n❌ Ошибка при генерации плана: {e}\n")
        sys.exit(1)


def parse_topics(raw_text, total_posts):
    topics = []
    for line in raw_text.strip().split("\n"):
        line = line.strip()
        if "|" in line:
            parts = line.split("|", 2)
            if len(parts) >= 2:
                post_type = parts[0].strip()
                theme = parts[1].strip()
                hashtags = parts[2].strip() if len(parts) > 2 else ""
                topics.append((post_type, theme, hashtags))
        if len(topics) >= total_posts:
            break
    return topics


def create_excel(topics, start_date, posts_per_week):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = date.today().strftime("%Y-%m-%d")
    filepath = os.path.join(OUTPUT_DIR, f"calendar_{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Контент-план"

    # Заголовки
    headers = ["Дата", "День недели", "Тип поста", "Тема", "Хэштеги", "Статус"]
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    # Чередующиеся цвета строк
    fill_even = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    status_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    type_colors = {
        "полезный": "27AE60",
        "история": "8E44AD",
        "кейс": "2980B9",
        "анонс": "E74C3C",
        "вирусный": "F39C12",
        "личное": "16A085",
    }

    # Генерируем даты по расписанию (равномерно в неделе)
    days_of_week = sorted(range(7), key=lambda x: x)
    step = max(1, 7 // posts_per_week)
    post_days = [i * step for i in range(posts_per_week)][:posts_per_week]

    current_date = start_date
    topic_idx = 0
    row_num = 2

    weeks = 2
    for week in range(weeks):
        week_start = start_date + timedelta(weeks=week)
        for day_offset in post_days:
            if topic_idx >= len(topics):
                break
            post_date = week_start + timedelta(days=day_offset)
            post_type, theme, hashtags = topics[topic_idx]
            day_name = DAYS_RU[post_date.weekday()]

            fill = fill_even if row_num % 2 == 0 else fill_odd

            values = [
                post_date.strftime("%d.%m.%Y"),
                day_name,
                post_type,
                theme,
                hashtags,
                "📝 Не начат",
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if col == 6:
                    cell.fill = status_fill
                else:
                    cell.fill = fill

                if col == 3:
                    color = "000000"
                    for key in type_colors:
                        if key in post_type.lower():
                            color = type_colors[key]
                            break
                    cell.font = Font(name="Arial", color=color, bold=True, size=10)
                else:
                    cell.font = Font(name="Arial", size=10)

            ws.row_dimensions[row_num].height = 45
            topic_idx += 1
            row_num += 1

    # Ширина колонок
    col_widths = [14, 14, 18, 45, 35, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Заголовок листа
    ws.insert_rows(1)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Контент-план на 2 недели | Старт: {start_date.strftime('%d.%m.%Y')}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color="2C3E50")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    ws.row_dimensions[1].height = 30

    wb.save(filepath)
    return filepath


def main():
    print("\n" + "─" * 50)
    print("   НейроПост v2.0 — Контент-план")
    print("─" * 50)

    passport = load_passport()

    try:
        posts_per_week = int(input("\nСколько постов в неделю? (1-7): ").strip())
        if not 1 <= posts_per_week <= 7:
            raise ValueError
    except ValueError:
        print("❌ Введи число от 1 до 7")
        sys.exit(1)

    start_input = input("Дата старта (ДД.ММ.ГГГГ, Enter = сегодня): ").strip()
    if start_input:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_input, "%d.%m.%Y").date()
        except ValueError:
            print("❌ Неверный формат даты. Используй ДД.ММ.ГГГГ")
            sys.exit(1)
    else:
        start_date = date.today()

    raw_topics = generate_topics(passport, posts_per_week)
    topics = parse_topics(raw_topics, posts_per_week * 2)

    filepath = create_excel(topics, start_date, posts_per_week)

    print("\n" + "─" * 50)
    print("✅ Контент-план готов!")
    print(f"📊 Файл: {filepath}")
    print(f"   {len(topics)} постов на 2 недели")
    print("─" * 50 + "\n")


if __name__ == "__main__":
    main()
